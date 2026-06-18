import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px
import pandas as pd
import io

# ตั้งค่า Layout ของหน้าเว็บ Streamlit ให้แสดงผลแบบเต็มหน้าจอ (Wide)
st.set_page_config(page_title="ระบบวิเคราะห์สัจพจน์ของคอลลาตซ์", layout="wide")

st.title("📊 แอปพลิเคชันอินเตอร์แอคทีฟ วิเคราะห์สัจพจน์ของคอลลาตซ์ (Collatz Conjecture)")
st.markdown("""
เว็บแอปพลิเคชันนี้จะคำนวณเส้นทางและลำดับคอลลาตซ์ (สมการ $3n + 1$) สำหรับทุกๆ จำนวนเต็มเริ่มต้นตั้งแตี่ 1 ไปจนถึงค่าสูงสุด $N$ ที่คุณกำหนด
* **🖥️ สำหรับคอมพิวเตอร์:** คลิกซ้ายค้างแล้วลากเพื่อ **เลื่อนกราฟ (Move/Pan)** และใช้ **ลูกกลิ้งเมาส์ (Mousewheel)** เพื่อซูมเข้า-ออก
* **📱 สำหรับมือถือ / แท็บเล็ต:** รูดนิ้วเดียวเพื่อเลื่อนกราฟ และใช้ **สองนิ้วกางออกหรือจีบเข้า (Pinch-to-zoom)** บนพื้นที่กราฟเพื่อซูมได้อย่างอิสระ
""")

# -------------------------------------------------------------------------
# 1. ส่วนควบคุมในแถบด้านข้าง (Sidebar)
# -------------------------------------------------------------------------
st.sidebar.header("⚙️ การตั้งค่าระบบ")
max_n = st.sidebar.number_input(
    "กำหนดค่า N สูงสุด", 
    min_value=1, 
    max_value=20000, 
    value=1000, 
    step=100,
    help="หมายเหตุ: การตั้งค่าที่สูงกว่า 5,000 อาจใช้เวลาในการประมวลผลไฟล์ Excel นนานขึ้นเล็กน้อยเนื่องจากต้องจัดฟอร์แมตข้อความจำนวนมากลงในเซลล์"
)

# -------------------------------------------------------------------------
# 2. ฟังก์ชันคำนวณประมวลผล (เปิดใช้งาน Cache เพื่อความรวดเร็ว)
# -------------------------------------------------------------------------
@st.cache_data
def run_collatz_simulation(limit_n):
    n_values = list(range(1, limit_n + 1))
    data_rows = []
    total_steps_list = []
    max_steps = 0

    for i in n_values:
        n = i
        steps_text = []
        step_count = 0
        while n != 1:
            prior = n
            if n % 2 == 0:
                n = n // 2
                steps_text.append(f"{prior}/2 = {n}")
            else:
                n = (prior * 3) + 1
                steps_text.append(f"3({prior})+1 = {n}")
            step_count += 1
                
        total_steps_list.append(step_count)
        if step_count > max_steps:
            max_steps = step_count
            
        data_rows.append({
            'n': i,
            'total_steps': step_count,
            'steps': steps_text
        })
    return n_values, total_steps_list, data_rows, max_steps

# เริ่มการคำนวณ
with st.spinner("กำลังคำนวณเส้นทางทางคณิตศาสตร์... กรุณารอสักครู่"):
    n_values, total_steps_list, data_rows, max_steps = run_collatz_simulation(max_n)

# -------------------------------------------------------------------------
# 3. การแสดงผลการคำนวณเฉลี่ยและสถิติสำคัญ (Metrics)
# -------------------------------------------------------------------------
avg_steps = sum(total_steps_list) / len(total_steps_list)
peak_steps = max(total_steps_list)
peak_n = n_values[total_steps_list.index(peak_steps)]

col1, col2, col3 = st.columns(3)
with col1:
    st.metric(label="📊 จำนวนขั้นตอนเฉลี่ย (Average Steps)", value=f"{avg_steps:.2f}")
with col2:
    st.metric(label="🚀 จำนวนขั้นตอนสูงสุดที่บันทึกได้ (Max Steps)", value=f"{peak_steps:,}")
with col3:
    st.metric(label="💎 ค่า N เริ่มต้นที่ทำให้เกิดขั้นตอนสูงสุด", value=f"{peak_n:,}")

st.markdown("---")

# -------------------------------------------------------------------------
# 4. กราฟอินเตอร์แอคทีฟ (ล็อกปุ่มซูมกรอบ เปลี่ยนเป็นโหมดเลื่อนกราฟ และเปิดทัชซูม)
# -------------------------------------------------------------------------
st.subheader("📈 แผนภูมิแสดงแนวโน้มจำนวนขั้นตอน (Interactive Trend Chart)")
st.caption("💡 ดับเบิ้ลคลิก (Double-click) บนพื้นที่กราฟเพื่อรีเซ็ตหน้าจอกลับสู่ค่าเริ่มต้น")

# แปลงข้อมูลเป็น DataFrame สำหรับ Plotly
chart_df = pd.DataFrame({
    "จำนวนเต็มเริ่มต้น (N)": n_values,
    "จำนวนขั้นตอนทั้งหมด": total_steps_list
})

fig = px.line(
    chart_df, 
    x="จำนวนเต็มเริ่มต้น (N)", 
    y="จำนวนขั้นตอนทั้งหมด",
    title=f"แผนภูมิสรุปจำนวนรอบการคำนวณจนกระทั่งเหลือ 1 (N = 1 ถึง {max_n:,})",
    labels={"จำนวนเต็มเริ่มต้น (N)": "ค่าเริ่มต้น (Starting N)", "จำนวนขั้นตอนทั้งหมด": "จำนวนขั้นตอนทั้งหมด (Total Steps)"},
    template="plotly_white"
)

# ปรับแต่งสีสันของเส้นกราฟ
fig.update_traces(
    line_color='#2F5597', 
    line_width=1.2 if max_n > 500 else 2.2,
    mode='lines+markers' if max_n <= 300 else 'lines',
    hovertemplate="<b>ค่าเริ่มต้น N:</b> %{x}<br><b>ขั้นตอนทั้งหมด:</b> %{y} รอบ<extra></extra>"
)

# ตั้งค่าให้เปลี่ยนโหมดจากตีกรอบซูมเป็นจับย้ายกราฟ (Pan) เป็นค่าหลักเริ่มต้น
fig.update_layout(
    title_font=dict(size=16, color='#1F4E78', family="Segoe UI"),
    hovermode="x unified",
    dragmode="pan"  # ล็อกโหมดคลิกลากเป็น "pan" เพื่อจับเคลื่อนย้ายกราฟไปรอบๆ
)

# ส่งพารามิเตอร์ config พิเศษไปยัง Streamlit
st.plotly_chart(
    fig, 
    use_container_width=True, 
    config={
        'scrollZoom': True,      # เปิดใช้งานซูมด้วย Mousewheel และ Pinch-to-zoom บนสมาร์ตโฟน/แท็บเล็ต
        'responsive': True,      # ปรับขนาดหน้าจอแบบยืดหยุ่นอัตโนมัติตามอุปกรณ์
        'modeBarButtonsToRemove': ['zoom2d', 'select2d', 'lasso2d']  # ลบเครื่องมือตีกรอบซูมออกเพื่อล็อกให้ใช้ระบบเลื่อนกราฟแทน
    }
)

st.markdown("---")

# -------------------------------------------------------------------------
# 5. ระบบส่งออกไฟล์ Excel ทางสถาปัตยกรรมหน่วยความจำ (In-Memory Buffer)
# -------------------------------------------------------------------------
st.subheader("📥 ดาวน์โหลดรายงานฉบับเต็ม (Excel Export)")
st.write("หากต้องการดูโครงสร้างข้อมูลดิบและการแจกแจงสูตรคำนวณอย่างละเอียดในแต่ละสเต็ป สามารถกดปุ่มสร้างไฟล์และดาวน์โหลดตารางข้อมูลด้านล่างนี้ได้เลยครับ")

def convert_data_to_excel_buffer(headers, data_rows, max_steps):
    wb = openpyxl.Workbook()
    
    ws_data = wb.active
    ws_data.title = "Collatz Analysis Data"
    ws_data.views.sheetView[0].showGridLines = True
    
    for col_idx, header in enumerate(headers, start=1):
        ws_data.cell(row=1, column=col_idx, value=header)
        
    for row_idx, row_data in enumerate(data_rows, start=2):
        ws_data.cell(row=row_idx, column=1, value=row_data['n'])
        ws_data.cell(row=row_idx, column=2, value=row_data['total_steps'])
        for step_idx, step_str in enumerate(row_data['steps'], start=3):
            ws_data.cell(row=row_idx, column=step_idx, value=step_str)
            
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10, color="000000")
    font_bold_data = Font(name="Segoe UI", size=10, bold=True, color="1F4E78")

    fill_header_n = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_header_steps = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    thin_border = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    for col_idx in range(1, len(headers) + 1):
        cell = ws_data.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header_n if col_idx <= 2 else fill_header_steps
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_cell

    for row_idx in range(2, len(data_rows) + 2):
        for col_idx in range(1, len(headers) + 1):
            cell = ws_data.cell(row=row_idx, column=col_idx)
            cell.border = border_cell
            if row_idx % 2 == 0:
                cell.fill = fill_zebra
                
            if col_idx <= 2:
                cell.font = font_bold_data
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = font_data
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws_data.column_dimensions['A'].width = 8
    ws_data.column_dimensions['B'].width = 12
    for col_idx in range(3, len(headers) + 1):
        ws_data.column_dimensions[get_column_letter(col_idx)].width = 16

    ws_data.freeze_panes = "C2"
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

headers = ["N", "Total Step"] + [f"step {i}" for i in range(1, max_steps + 1)]

if st.button("🔄 เริ่มต้นจัดทำไฟล์ Excel สำหรับดาวน์โหลด"):
    with st.spinner("ระบบกำลังรวบรวมข้อมูลสเต็ปและตกแต่งสไตล์ Excel..."):
        excel_data = convert_data_to_excel_buffer(headers, data_rows, max_steps)
        
        st.download_button(
            label="💾 คลิกเพื่อดาวน์โหลดไฟล์รายงานฉบับสมบูรณ์ (.xlsx)",
            data=excel_data,
            file_name=f"collatz_conjecture_analysis_1_to_{max_n}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )